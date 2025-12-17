import os
import time
import logging
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Iterable, Tuple
import azure.functions as func
import pyodbc
import openpyxl

# ========= CONFIG =========
# Read from Function App settings (Configuration → Application settings)
SQL_SERVER = "purenvqld.database.windows.net"
SQL_DATABASE = "LaboratoryQLD"
SQL_USERNAME = "reportabledatadownloader"
SQL_PASSWORD = "Rep0r7D47aD0wn"

# Map HTML group ids → fully-qualified table names (include schema!)
GROUP_TO_TABLE: Dict[str, str] = {
    "externalds": "[Jackson].[Dust Suppression]",     # External Analytes
    "ccp": "[Jackson].[DMS]",
    "treff": "[Jackson].[TreatedEffluent]",
    "lsfw": "[Jackson].[LandfillSWDischarge]",
    "swall": "[Jackson].[Stormwater]",
    "envin": "[Jackson.[Environmental Creek]",
    "bior": "[Jackson].[BioRemediation Testing]"
}

# Whitelist analyte columns per table to avoid SQL injection via column names.
# ⚠️ 
ALLOWED_COLUMNS: Dict[str, Iterable[str]] = {
    "[Jackson].[Dust Suppression]": {
        "File","Sample Location","Sample Date","Arsenic","Beryllium","Cadmium","Chromium","Copper","Cobalt","Nickel","Lead","Zinc","Manganese","Selenium","Silver","Vanadium","Boron","Mercury","Total Organic Carbon","C10 - C14 Fraction","C15 - C28 Fraction"
        ,"C29 - C36 Fraction","C10 - C36 Fraction (sum)",">C10 - C16 Fraction",">C16 - C34 Fraction",">C34 - C40 Fraction",">C10 - C40 Fraction (sum)",">C10 - C16 Fraction minus Naphthalene (F2)","Phenol","2-Chlorophenol","2-Methylphenol","3- & 4-Methylphenol"
        ,"2-Nitrophenol","2.4-Dimethylphenol","2.4-Dichlorophenol","2.6-Dichlorophenol","4-Chloro-3-methylphenol","2.4.6-Trichlorophenol","2.4.5-Trichlorophenol","Pentachlorophenol","Sum of Phenols","C6 - C9 Fraction","C6 - C10 Fraction","C6 - C10 Fraction minus BTEX (F1)"
        ,"Benzene","Toluene","Ethylbenzene","meta- & para-Xylene","ortho-Xylene","Total Xylenes","Sum of BTEX","Naphthalene","Escherichia coli","Sodium","Potassium","Magnesium","Calcium","Perfluorobutane sulfonic acid","Perfluoropentane sulfonic acid","Perfluorohexane sulfonic acid"
        ,"Perfluoroheptane sulfonic acid","Perfluorooctane sulfonic acid","Perfluorodecane sulfonic acid","Perfluorobutanoic acid","Perfluoropentanoic acid","Perfluorohexanoic acid","Perfluoroheptanoic","Perfluorooctanoic acid","Perfluorononanoic acid","Perfluorodecanoic acid","Perfluoroundecanoic acid"
        ,"Perfluorododecanoic acid","Perfluorotridecanoic acid","Perfluorotetradecanoic acid","Perfluorooctane sulfonamide","N-Methyl perfluorooctane sulfonamide","N-Ethyl perfluorooctane sulfonamide","N-Methyl perfluorooctane sulfonamidoethanol","N-Ethyl perfluorooctane sulfonamidoethanol","N-Methyl perfluorooctane sulfonamidoacetic acid"
        ,"N-Ethyl perfluorooctane sulfonamidoacetic acid","4:2 Fluorotelomer sulfonic acid","6:2 Fluorotelomer sulfonic acid","8:2 Fluorotelomer sulfonic acid","10:2 Fluorotelomer sulfonic acid","Sum of PFAS","Sum of PFHxS and PFOS","Sum of TOP C4 - C14 Carboxylates and C4-C8 Sulfonates","Sum of TOP C4 - C14 as Fluorine","Total Organic Fluorine"
    },
    "[Jackson].[DMS]": {
        "File","Sample Date","Sample Location","pH Value","Total Soluble Salts","Moisture Content","Extractable Boron","Calcium","Magnesium","Sodium","Potassium","Arsenic","Barium","Boron","Cadmium","Chromium","Copper","Lead","Nickel","Selenium","Silver","Vanadium","Zinc","ICP-MS Arsenic","Mercury","Hexavalent Chromium","Trivalent Chromium"
        ,"C10 - C14 Fraction","C15 - C28 Fraction","C29 - C36 Fraction","C10 - C36 Fraction (sum)",">C10 - C16 Fraction",">C16 - C34 Fraction",">C34 - C40 Fraction",">C10 - C40 Fraction (sum)",">C10 - C16 Fraction minus Naphthalene (F2)","Phenol","2-Chlorophenol","2-Methylphenol","3- & 4-Methylphenol","2-Nitrophenol"
        ,"2.4-Dimethylphenol","2.4-Dichlorophenol","2.6-Dichlorophenol","4-Chloro-3-methylphenol","2.4.6-Trichlorophenol","2.4.5-Trichlorophenol","Pentachlorophenol","PAH Naphthalene","Acenaphthylene","Acenaphthene","Fluorene","Phenanthrene","Anthracene","Fluoranthene","Pyrene","Benz(a)anthracene","Chrysene","Benzo(b+j)fluoranthene"
        ,"Benzo(k)fluoranthene","Benzo(a)pyrene","Indeno(1.2.3.cd)pyrene","Dibenz(a.h)anthracene","Benzo(g.h.i)perylene","Benzo(a)pyrene TEQ (zero)","Benzo(a)pyrene TEQ (half LOR)","Benzo(a)pyrene TEQ (LOR)","Sum of polycyclic aromatic hydrocarbons","C6 - C10 Fraction","C6 - C10 Fraction minus BTEX (F1)","C6 - C9 Fraction"
        ,"Benzene","Toluene","Ethylbenzene","meta- & para-Xylene","ortho-Xylene","Total Xylenes","Sum of BTEX","Naphthalene"
    },
    "[Jackson].[TreatedEffluent]":{
        "pH Value","Total Dissolved Solids @180°C","Nitrite + Nitrate as N","Total Kjeldahl Nitrogen as N","Total Nitrogen As N","Total Phosphorus As P","Total Organic Carbon","Biochemical Oxygen Demand","Escherichia coli"
    },
    "[Jackson].[Environmental Creek]": {
        "File","Sample Date","Sample Location","pH","Electrical Conductivity","Total Dissolved Solids","Dissolved Oxygen","Suspended Solids (SS)","Total Organic Carbon","Dissolved Arsenic","Dissolved Cadmium","Dissolved Chromium","Dissolved Copper","Dissolved Lead","Dissolved Nickel","Dissolved Zinc","Dissolved Mercury","Arsenic","Cadmium"
        ,"Chromium","Copper","Lead","Nickel","Zinc","Mercury","Perfluorobutane sulfonic acid (PFBS)","Perfluoropentane sulfonic acid (PFPeS)","Perfluorohexane sulfonic acid (PFHxS)","Perfluoroheptane sulfonic acid (PFHpS)","Perfluorooctane sulfonic acid (PFOS)","Perfluorodecane sulfonic acid (PFDS)","Perfluorobutanoic acid (PFBA)"
        ,"Perfluoropentanoic acid (PFPeA)","Perfluorohexanoic acid (PFHxA)","Perfluoroheptanoic acid (PFHpA)","Perfluorooctanoic acid (PFOA)","Perfluorononanoic acid (PFNA)","Perfluorodecanoic acid (PFDA)","Perfluoroundecanoic acid (PFUnDA)","Perfluorododecanoic acid (PFDoDA)","Perfluorotridecanoic acid (PFTrDA)","Perfluorotetradecanoic acid (PFTeDA)"
        ,"Perfluorooctane sulfonamide (FOSA)","N-Methyl perfluorooctane sulfonamide (MeFOSA)","N-Ethyl perfluorooctane sulfonamide (EtFOSA)","N-Methyl perfluorooctane sulfonamidoethanol (MeFOSE)","N-Ethyl perfluorooctane sulfonamidoethanol (EtFOSE)","N-Methyl perfluorooctane sulfonamidoacetic acid (MeFOSAA)","N-Ethyl perfluorooctane sulfonamidoacetic acid (EtFOSAA)"
        ,"4:2 Fluorotelomer sulfonic acid (4:2 FTS)","6:2 Fluorotelomer sulfonic acid (6:2 FTS)","8:2 Fluorotelomer sulfonic acid (8:2 FTS)","10:2 Fluorotelomer sulfonic acid (10:2 FTS)","Sum of PFAS","Sum of PFHxS and PFOS","Sum of PFAS (WA DER List)","Total Organic Fluorine"
    },
    "[Jackson].[LandfillSWDischarge]": {
        "pHR1","TDSR1","ECR1","DOR1","TSSR1","TOCR1","pHR2","TDSR2","ECR2","DOR2","TSSR2","TOCR2"
    },
    "[Jackson].[Stormwater]": {
        "File","Sample Date","Sample Location","pH Value","Total Dissolved Solids @180°C","Calcium","Magnesium","Sodium","Potassium","Arsenic","Beryllium","Barium","Cadmium","Chromium","Cobalt","Copper","Lead","Manganese","Nickel","Selenium","Silver","Vanadium","Zinc","Boron","Mercury","Trivalent Chromium","Hexavalent Chromium","C10 - C14 Fraction"
        ,"C15 - C28 Fraction","C29 - C36 Fraction","C10 - C36 Fraction (sum)",">C10 - C16 Fraction",">C16 - C34 Fraction",">C34 - C40 Fraction",">C10 - C40 Fraction (sum)",">C10 - C16 Fraction minus Naphthalene (F2)","Phenol","2-Chlorophenol","2-Methylphenol","3- & 4-Methylphenol","2-Nitrophenol","2.4-Dimethylphenol"
        ,"2.4-Dichlorophenol","2.6-Dichlorophenol","4-Chloro-3-methylphenol","2.4.6-Trichlorophenol","2.4.5-Trichlorophenol","Pentachlorophenol","PAH Naphthalene","Acenaphthylene","Acenaphthene","Fluorene","Phenanthrene","Anthracene","Fluoranthene","Pyrene","Benz(a)anthracene","Chrysene","Benzo(b+j)fluoranthene","Benzo(k)fluoranthene"
        ,"Benzo(a)pyrene","Indeno(1.2.3.cd)pyrene","Dibenz(a.h)anthracene","Benzo(g.h.i)perylene","Sum of polycyclic aromatic hydrocarbons","Benzo(a)pyrene TEQ (zero)","C6 - C9 Fraction","C6 - C10 Fraction","C6 - C10 Fraction  minus BTEX (F1)","Benzene","Toluene","Ethylbenzene","meta- & para-Xylene","ortho-Xylene","Total Xylenes"
        ,"Sum of BTEX","Naphthalene"
    },
    "[Jackson].[BioRemediation Testing]": {
        "File","Sample Date","Sample Name","Moisture Content","Nitrite + Nitrate as N (Sol.)","Total Kjeldahl Nitrogen as N","Total Nitrogen as N","Total Organic Carbon","Thermotolerant Coliforms","E.coli by MPN"
    },
    "[Jackson].[Pump Well]": {
        "File","Sample Date","Sample Name","Perfluorobutane sulfonic acid (PFBS)","Perfluoropentane sulfonic acid (PFPeS)","Perfluorohexane sulfonic acid (PFHxS)","Perfluoroheptane sulfonic acid (PFHpS)","Perfluorooctane sulfonic acid (PFOS)","Perfluorodecane sulfonic acid (PFDS)","Perfluorobutanoic acid (PFBA)","Perfluoropentanoic acid (PFPeA)"
        ,"Perfluorohexanoic acid (PFHxA)""Perfluoroheptanoic acid (PFHpA)","Perfluorooctanoic acid (PFOA)","Perfluorononanoic acid (PFNA)","Perfluorodecanoic acid (PFDA)","Perfluoroundecanoic acid (PFUnDA)","Perfluorododecanoic acid (PFDoDA)","Perfluorotridecanoic acid (PFTrDA)","Perfluorotetradecanoic acid (PFTeDA)","Perfluorooctane sulfonamide (FOSA)"
        ,"N-Methyl perfluorooctane sulfonamide (MeFOSA)","N-Ethyl perfluorooctane sulfonamide (EtFOSA)","N-Methyl perfluorooctane sulfonamidoethanol (MeFOSE)","N-Ethyl perfluorooctane sulfonamidoethanol (EtFOSE)","N-Methyl perfluorooctane sulfonamidoacetic acid (MeFOSAA)","N-Ethyl perfluorooctane sulfonamidoacetic acid (EtFOSAA)","4:2 Fluorotelomer sulfonic acid (4:2 FTS)"
        ,"6:2 Fluorotelomer sulfonic acid (6:2 FTS)","8:2 Fluorotelomer sulfonic acid (8:2 FTS)","10:2 Fluorotelomer sulfonic acid (10:2 FTS)","Sum of PFAS","Sum of PFHxS and PFOS","Sum of PFAS (WA DER List)"
    }
}

ID_COLUMNS = ["File","Sample Date", "Sample Location", "Sample Name"]

# Non-analyte identifier columns you always want back
TABLE_ID_COLUMNS: Dict[str, List[str]] = {
    "[Jackson].[Dust Suppression]": ["File","Sample Date", "Sample Location"],  # doesn’t have Sample Location
    "[Jackson].[DMS]": ["File","Sample Date", "Sample Location"],
    "[Jackson].[TreatedEffluent]": ["File", "Sample Date"],
    "[Jackson].[Environmental Creek]": ["File", "Sample Date", "Sample Location"],
    "[Jackson].[LandfillSWDischarge]": ["Sample Date", "DateR2"],
    "[Jackson].[Stormwater]": ["File","Sample Date", "Sample Location"],
    "[Jackson].[BioRemediation Testing]": ["File","Sample Date", "Sample Name"],
    "[Jackson].[Pump Well]": ["File","Sample Date", "Sample Name"]
}

# ========= DB CONNECT =========
def connect_with_fallback(timeout_seconds: int = 60) -> pyodbc.Connection:
    """
    Try ODBC Driver 18 then 17. Increase Connection Timeout and retry a few times
    (useful if Azure SQL Serverless is resuming).
    """
    drivers = ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"]
    last_exc = None

    for driver in drivers:
        conn_str = (
            f"Driver={{{driver}}};"
            f"Server=tcp:{SQL_SERVER},1433;"
            f"Database={SQL_DATABASE};"
            f"Uid={SQL_USERNAME};"
            f"Pwd={SQL_PASSWORD};"
            "Encrypt=yes;"
            "TrustServerCertificate=no;"
            f"Connection Timeout={timeout_seconds};"
        )
        for attempt in range(3):
            try:
                return pyodbc.connect(conn_str)
            except Exception as e:
                last_exc = e
                logging.warning(f"Connect attempt {attempt+1}/3 with {driver} failed: {e}")
                time.sleep(3)
    # If we get here, all attempts failed
    raise last_exc

# ========= HELPERS =========
def normalize_payload(data) -> Dict[str, List[str]]:
    """
    Support two payload shapes:
      A) { "selections": { "external": ["arsenic", ...], "pfastopa": ["pfos"] }, "startDate": "...", "endDate": "..." }
      B) { "selections": [ { "table": "Jackson.DSExt", "analyte":"arsenic" }, ... ], ... }
    Returns a dict { group_key_or_table: [analytes...] } keyed by group-id (preferred) if possible.
    """
    sel = data.get("selections", [])
    if isinstance(sel, dict):
        # Already grouped by HTML group id
        return {k: list(v) for k, v in sel.items() if v}
    elif isinstance(sel, list):
        grouped: Dict[str, List[str]] = {}
        for item in sel:
            table = item.get("table")
            analyte = item.get("analyte")
            if not table or not analyte:
                continue
            # Try to reverse-map table back to group id; if not found, use the table name as the key
            group_key = None
            for g, t in GROUP_TO_TABLE.items():
                if t.lower() == table.lower():
                    group_key = g
                    break
            key = group_key if group_key else table
            grouped.setdefault(key, []).append(analyte)
        return grouped
    else:
        return {}

def whitelist_columns(table: str, requested: Iterable[str]) -> List[str]:
    allowed = set(ALLOWED_COLUMNS.get(table, []))
    return [c for c in requested if c in allowed]

def build_select_sql(table: str, analyte_cols: List[str]) -> str:
    # Fall back to global ID_COLUMNS if table not in map
    id_cols = TABLE_ID_COLUMNS.get(table, ID_COLUMNS)

    selected_cols = id_cols + analyte_cols
    cols_sql = ", ".join(f"[{c}]" for c in selected_cols)  # bracket-quote identifiers
    return f"SELECT {cols_sql} FROM {table} WHERE [Sample Date] BETWEEN ? AND ?"
    
def safe_sheet_name(name: str) -> str:
    # Excel sheet name: max 31 chars, no []:*?/\
    bad = '[]:*?/\\'
    for ch in bad:
        name = name.replace(ch, "-")
    return (name or "Sheet")[:31]

# ========= MAIN FUNCTION =========

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Processing /download-excel request")
    try:
        data = req.get_json()
    except ValueError:
        return func.HttpResponse("Invalid JSON body.", status_code=400)

    start_date = data.get("startDate")
    end_date = data.get("endDate")
    if not start_date or not end_date:
        return func.HttpResponse("Both startDate and endDate are required.", status_code=400)

    grouped = normalize_payload(data)
    if not grouped:
        return func.HttpResponse("No analytes selected.", status_code=400)

    # Open workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)



    try:
        

        conn = connect_with_fallback(timeout_seconds=60)
        cursor = conn.cursor()

        any_rows_written = False

        for group_key, analytes in grouped.items():
            table = GROUP_TO_TABLE.get(group_key, group_key)
            if table not in ALLOWED_COLUMNS:
                logging.warning(f"Skipping unknown/unauthorized table: {table}")
                continue

            analyte_cols = whitelist_columns(table, analytes)
            if not analyte_cols:
                logging.info(f"No valid analyte columns for {table}, requested: {analytes}")
                continue

            sql = build_select_sql(table, analyte_cols)
            logging.info(f"Running query for {table}: {sql}")
            cursor.execute(sql, (start_date, end_date))
            rows = cursor.fetchall()
            columns = [d[0] for d in cursor.description]

            logging.info(f"Grouped selections: {grouped}")
            logging.info(f"SQL about to run: {sql}")
            logging.info(f"Row count returned: {len(rows)}")


            ws = wb.create_sheet(title=safe_sheet_name(group_key))
            ws.append(columns)

            if rows:
                for row in rows:
                    ws.append(list(row))
                any_rows_written = True
            else:
                ws.append(["No data found for this selection."])

        if not wb.worksheets:  # safety: if nothing created
            ws = wb.create_sheet(title="Results")
            ws.append(["No data found at all."])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return func.HttpResponse(
            body=output.getvalue(),
            status_code=200,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logging.error(f"Download error: {e}", exc_info=True)
        return func.HttpResponse(f"Error: {e}", status_code=500)
    finally:
        try: cursor.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass